#!/usr/bin/env python3
"""
assemble_xlsx.py — turn one or more trials TSVs (from extract_target_trials.py) into a
styled multi-sheet .xlsx:  README, Drug_Summary, Trials (main).

Usage:
  python3 assemble_xlsx.py --trials trials_rows.tsv --out database.xlsx \
      --title "My target -> drug -> trial repurposing DB"
  # multiple TSVs (e.g. batches) are concatenated + de-duplicated on (drug, NCT):
  python3 assemble_xlsx.py --trials a.tsv b.tsv --out database.xlsx
"""
import argparse, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN_COLS=["targets","drug","nct","phase","status","start_date","enrollment","conditions",
         "n_primary","n_secondary","primary_outcomes","secondary_outcomes","biomarker"]
PDKEY=("parkinson","neurodeg","dementia","alzheimer","cognit","dopamin","lewy","tremor","synuclein")

HDR=PatternFill("solid",fgColor="1F4E5F"); HF=Font(color="FFFFFF",bold=True,size=11)
BIO=PatternFill("solid",fgColor="FFF2CC")
thin=Side(style="thin",color="D0D0D0"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")

def load(paths):
    recs=[]
    for p in paths:
        for ln in open(p).read().splitlines()[1:]:
            if not ln.strip(): continue
            c=ln.split("\t")
            if len(c)<13: continue
            recs.append(dict(zip(IN_COLS,c[:13])))
    seen=set(); uniq=[]
    for r in recs:
        k=(r["drug"].upper(),r["nct"])
        if k in seen: continue
        seen.add(k); uniq.append(r)
    uniq.sort(key=lambda r:(r["drug"].upper(),r["nct"]))
    return uniq

def style_header(ws,n):
    for j in range(1,n+1):
        c=ws.cell(1,j); c.fill=HDR; c.font=HF
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=26
    ws.auto_filter.ref=f"A1:{get_column_letter(n)}{ws.max_row}"

def setw(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--trials",nargs="+",required=True)
    ap.add_argument("--out",required=True)
    ap.add_argument("--title",default="Target -> Approved-Drug -> Clinical-Trial repurposing database")
    a=ap.parse_args()
    recs=load(a.trials)
    wb=Workbook()

    ws=wb.active; ws.title="README"
    txt=[(a.title,True,14),("",False,11),
     ("Pipeline: gene targets -> human UniProt -> approved drugs (ChEMBL max_phase=4) ->",False,11),
     ("ALL ClinicalTrials.gov trials for those drugs (every indication) + outcome/biomarker measures.",False,11),
     ("Built with Paperclip (ChEMBL + AACT/ClinicalTrials.gov).",False,11),("",False,11),
     ("CAVEATS",True,12),
     ("- Scope = ALL indications; most trials are the drug's on-label use, NOT the disease of interest.",False,11),
     ("  Filter the 'Conditions' column for your indication.",False,11),
     ("- Target->drug = annotated mechanism, not proof of efficacy for repurposing. Hypothesis-generating.",False,11),
     ("- Outcome text is reassembled from registry 'measure' fields, capped (~220 chars); N_primary/",False,11),
     ("  N_secondary hold true counts, and the NCT link has the full record.",False,11),
     ("- Biomarker_measures=Yes if any outcome matched a biomarker keyword (see extract_target_trials.py).",False,11),
     ("  It is a screening aid — verify per trial.",False,11)]
    for i,(t,b,s) in enumerate(txt,1):
        ws.cell(i,1,t).font=Font(bold=b,size=s,color=("1F4E5F" if b and s>=12 else "000000"))
    ws.column_dimensions["A"].width=100

    ws=wb.create_sheet("Drug_Summary")
    ws.append(["Drug","Targets","N_trials","N_biomarker_trials","N_indication_relevant_trials(PD/neuro)"])
    by={}
    for r in recs: by.setdefault(r["drug"],[]).append(r)
    for drug in sorted(by,key=str.upper):
        rs=by[drug]
        ws.append([drug, rs[0]["targets"], len(rs),
            sum(1 for r in rs if r["biomarker"]=="Yes"),
            sum(1 for r in rs if any(k in r["conditions"].lower() for k in PDKEY))])
    style_header(ws,5); setw(ws,[24,28,10,18,34])
    for r in range(2,ws.max_row+1): ws.cell(r,2).alignment=WRAP

    ws=wb.create_sheet("Trials")
    hdr=["Targets","Drug","NCT_ID","URL","Phase","Status","Start_date","Enrollment","Conditions",
         "N_primary","N_secondary","Primary_outcome_measures","Secondary_outcome_measures","Biomarker_measures"]
    ws.append(hdr)
    for r in recs:
        url=f"https://clinicaltrials.gov/study/{r['nct']}" if r["nct"] else ""
        ws.append([r["targets"],r["drug"],r["nct"],url,r["phase"],r["status"],r["start_date"],
            r["enrollment"],r["conditions"],r["n_primary"],r["n_secondary"],
            r["primary_outcomes"],r["secondary_outcomes"],r["biomarker"]])
    style_header(ws,len(hdr)); setw(ws,[22,20,13,34,12,14,11,10,34,9,9,50,50,14])
    for row in range(2,ws.max_row+1):
        for col in (1,9,12,13): ws.cell(row,col).alignment=WRAP
        if ws.cell(row,14).value=="Yes":
            ws.cell(row,14).fill=BIO; ws.cell(row,14).font=Font(bold=True,color="7F6000")
        u=ws.cell(row,4).value
        if u: ws.cell(row,4).hyperlink=u; ws.cell(row,4).font=Font(color="0563C1",underline="single")

    wb.save(a.out)
    print(f"wrote {a.out}: {len(by)} drugs, {len(recs)} trial rows, "
          f"{sum(1 for r in recs if r['biomarker']=='Yes')} biomarker trials")

if __name__=="__main__": main()
